from pathlib import Path
from app_config import *
import json, os, traceback
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from json_repair import repair_json
import json_repair

ref_json = BASE_PATH+'JSON.txt'

def flatten_json(data_file):
    """
    Flattens a nested dictionary so that all keys at n-1 depth are promoted to the top level,
    and their parent keys become empty dicts. This works for arbitrary depth n.
    Example:
    Input: {"Income": {"Revenue": {...}}}
    Output: {"Income": {}, "Revenue": {...}}
    """
    with open(data_file, "r", encoding="utf-8") as f:
        #data_json = json.load(f)
        data_json = json_repair.load(f)
    
    if not isinstance(data_json, dict):
        return data_json
    result = {}
    def _flatten(obj, parent=None):
        if isinstance(obj, dict):
            for k, v in obj.items():
                if isinstance(v, dict):
                    # Promote children to top level
                    for child_k, child_v in v.items():
                        #result[child_k] = child_v
                        if not isinstance(child_v, dict):
                            #last value is final value so the parent dict should be final
                            result[k] = v
                        else:
                            result[k] = {}
                            _flatten(v)
                        break
                else:
                    result[k] = v
        else:
            if parent:
                result[parent] = obj
    _flatten(data_json)

    #flat_file = data_file+"_flat.json"
    #with open(flat_file, "w", encoding="utf-8") as f:
    #    json.dump(result, f, indent=4, ensure_ascii=False)
    
    return result


def json_to_xlsx_with_formulas(data_json_path, json_tmpl_path, excel_path):
    global last_val
    global last_col_title 
    flat_json = flatten_json(data_json_path)

    with open(json_tmpl_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()
    # Remove default sheet if present
    if "Sheet" in wb.sheetnames and len(data) > 0:
        std = wb["Sheet"]
        wb.remove(std)

    for sheet_name, rows in data.items():
        ws = wb.create_sheet(title=sheet_name)
        # Collect all columns to determine header order
        all_cols = set()
        for row in rows.values():
            for k in row.keys():
                if k.endswith('_value') or k.endswith('_formula'):
                    all_cols.add(k.split('_')[0])
        col_letters = sorted(all_cols)
        # Write header
        for idx, col in enumerate(col_letters, 1):
            ws.cell(row=1, column=idx, value=col)
        # Write data
        for row_idx, (row_key, row_data) in enumerate(rows.items(), start=1):
            last_col_title = ""
            for col in col_letters:
                cell = ws.cell(row=row_idx, column=col_letters.index(col)+1)
                value_key = f"{col}_value"
                formula_key = f"{col}_formula"
                # Set value or formula
                if formula_key in row_data:
                    cell.value = row_data[formula_key]
                    if not str(cell.value).startswith("="):
                        cell.value = "=" + str(cell.value)
                elif value_key in row_data:
                    cell.value = row_data[value_key]
                    last_col_title = cell.value

            if(sheet_name=="Cash Flow"): 
                continue # Skip for Cash Flow sheet as it is all formnula driven
            if(last_col_title!="" and flat_json.__contains__(last_col_title)):
                value_dict = flat_json[last_col_title]
                if(not isinstance(value_dict, dict)):
                    #print(f"Invalid value: {last_col_title}={value_dict}")
                    continue
                col='B'
                idx=1
                for k,v in value_dict.items():
                    idx+=1
                    ws.cell(row=1, column=col_letters.index(col)+idx, value=k)
                    last_val = v
                    ws.cell(row=row_idx, column=col_letters.index(col)+idx, value=v)
    wb.save(excel_path)

def json_to_xlsx(data_json, xl_template, xl_output):
    json_to_xlsx_with_formulas(data_json, xl_template+".json", xl_output)

